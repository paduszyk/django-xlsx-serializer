from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.db import IntegrityError

from xlsx_serializer.core import Deserializer

from tests.models import BlankFieldModel, DummyModel, NullFieldModel

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.parametrize(
    "sheet_name",
    [
        "tests.DummyModel",
        "tests.dummymodel",
        "DummyModel",
        "dummymodel",
    ],
    ids=[
        "label",
        "label_lower",
        "model_name",
        "model_name_lower",
    ],
)
def test_deserializer_identifies_model_by_sheet_name(
    fixture_path: Path,
    sheet_name: str,
) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(sheet_name)
    worksheet["A1"].value = "id"
    worksheet["A2"].value = 1
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, DummyModel)
    assert deserialized_object.object.pk == 1


def test_deserializer_ignores_non_model_fields(fixture_path: Path) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.DummyModel")
    worksheet["A1"].value = "id"
    worksheet["A2"].value = 1
    worksheet["B1"].value = "not_a_field"
    worksheet["B2"].value = None
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, DummyModel)
    assert deserialized_object.object.pk == 1


def test_deserializer_ignores_empty_left_columns(fixture_path: Path) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.DummyModel")
    worksheet["B1"].value = "id"
    worksheet["B2"].value = 1
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, DummyModel)
    assert deserialized_object.object.pk == 1


def test_deserializer_ignores_empty_top_rows(fixture_path: Path) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.DummyModel")
    worksheet["A2"].value = "id"
    worksheet["A3"].value = 1
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, DummyModel)
    assert deserialized_object.object.pk == 1


def test_deserializer_reads_empty_cells_as_blank(fixture_path: Path) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.BlankFieldModel")
    worksheet["A1"].value = "id"
    worksheet["A2"].value = 1
    worksheet["B1"].value = "blank_field"
    worksheet["B2"].value = None
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, BlankFieldModel)
    assert deserialized_object.object.blank_field == ""


def test_deserializer_reads_empty_cells_as_none(fixture_path: Path) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.NullFieldModel")
    worksheet["A1"].value = "id"
    worksheet["A2"].value = 1
    worksheet["B1"].value = "null_field"
    worksheet["B2"].value = None
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    assert isinstance(deserialized_object.object, NullFieldModel)
    assert deserialized_object.object.null_field is None


@pytest.mark.parametrize(
    "empty_value",
    [
        "",
        None,
    ],
    ids=[
        "empty string",
        "None",
    ],
)
@pytest.mark.django_db()
def test_django_handles_empty_non_nullable_fields(
    fixture_path: Path,
    empty_value: str | None,
) -> None:
    # Arrange.
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("tests.NotNullFieldModel")
    worksheet["A1"].value = "id"
    worksheet["A2"].value = 1
    worksheet["B1"].value = "not_null_field"
    worksheet["B2"].value = empty_value
    workbook.save(fixture_path)

    # Act.
    deserializer = iter(Deserializer(fixture_path))
    deserialized_object = next(deserializer)

    # Assert.
    with pytest.raises(IntegrityError):
        deserialized_object.save()
