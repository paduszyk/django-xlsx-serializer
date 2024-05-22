from __future__ import annotations

from typing import TYPE_CHECKING
from unittest import mock

import openpyxl
import pytest

from django.core.management import call_command

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db()
def test_loaddata_command_invokes_deserializer(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DummyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    wb.save(fixture_path)

    # Act.
    with mock.patch("xlsx_serializer.Deserializer") as deserializer_mock:
        call_command("loaddata", fixture_path)

    # Assert.
    deserializer_mock.assert_called()
