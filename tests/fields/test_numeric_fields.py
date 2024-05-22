from __future__ import annotations

from decimal import Decimal
from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import (
    BigIntegerFieldModel,
    DecimalFieldModel,
    FloatFieldModel,
    IntegerFieldModel,
    PositiveBigIntegerFieldModel,
    PositiveIntegerFieldModel,
    PositiveSmallIntegerFieldModel,
    SmallIntegerFieldModel,
)

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db()
def test_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.IntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = IntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.integer_field == 42


@pytest.mark.django_db()
def test_integer_field_is_serialized() -> None:
    # Arrange.
    obj = IntegerFieldModel._default_manager.create(pk=1, integer_field=42)

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.IntegerFieldModel"]["A1"].value == "id"
    assert wb["tests.IntegerFieldModel"]["A2"].value == 1
    assert wb["tests.IntegerFieldModel"]["B1"].value == "integer_field"
    assert wb["tests.IntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_big_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.BigIntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "big_integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = BigIntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.big_integer_field == 42


@pytest.mark.django_db()
def test_big_integer_field_is_serialized() -> None:
    # Arrange.
    obj = BigIntegerFieldModel._default_manager.create(pk=1, big_integer_field=42)

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.BigIntegerFieldModel"]["A1"].value == "id"
    assert wb["tests.BigIntegerFieldModel"]["A2"].value == 1
    assert wb["tests.BigIntegerFieldModel"]["B1"].value == "big_integer_field"
    assert wb["tests.BigIntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_small_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.SmallIntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "small_integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = SmallIntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.small_integer_field == 42


@pytest.mark.django_db()
def test_small_integer_field_is_serialized() -> None:
    # Arrange.
    obj = SmallIntegerFieldModel._default_manager.create(pk=1, small_integer_field=42)

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.SmallIntegerFieldModel"]["A1"].value == "id"
    assert wb["tests.SmallIntegerFieldModel"]["A2"].value == 1
    assert wb["tests.SmallIntegerFieldModel"]["B1"].value == "small_integer_field"
    assert wb["tests.SmallIntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_positive_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PositiveIntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "positive_integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = PositiveIntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.positive_integer_field == 42


@pytest.mark.django_db()
def test_positive_integer_field_is_serialized() -> None:
    # Arrange.
    obj = PositiveIntegerFieldModel._default_manager.create(
        pk=1,
        positive_integer_field=42,
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.PositiveIntegerFieldModel"]["A1"].value == "id"
    assert wb["tests.PositiveIntegerFieldModel"]["A2"].value == 1
    assert wb["tests.PositiveIntegerFieldModel"]["B1"].value == "positive_integer_field"
    assert wb["tests.PositiveIntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_positive_big_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("PositiveBigIntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "positive_big_integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = PositiveBigIntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.positive_big_integer_field == 42


@pytest.mark.django_db()
def test_positive_big_integer_field_is_serialized() -> None:
    # Arrange.
    obj = PositiveBigIntegerFieldModel._default_manager.create(
        pk=1,
        positive_big_integer_field=42,
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["PositiveBigIntegerFieldModel"]["A1"].value == "id"
    assert wb["PositiveBigIntegerFieldModel"]["A2"].value == 1
    assert wb["PositiveBigIntegerFieldModel"]["B1"].value == "positive_big_integer_field"  # fmt: skip
    assert wb["PositiveBigIntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_positive_small_integer_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("PositiveSmallIntegerFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "positive_small_integer_field"
    ws["B2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = PositiveSmallIntegerFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.positive_small_integer_field == 42


@pytest.mark.django_db()
def test_positive_small_integer_field_is_serialized() -> None:
    # Arrange.
    obj = PositiveSmallIntegerFieldModel._default_manager.create(
        pk=1,
        positive_small_integer_field=42,
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["PositiveSmallIntegerFieldModel"]["A1"].value == "id"
    assert wb["PositiveSmallIntegerFieldModel"]["A2"].value == 1
    assert wb["PositiveSmallIntegerFieldModel"]["B1"].value == "positive_small_integer_field"  # fmt: skip
    assert wb["PositiveSmallIntegerFieldModel"]["B2"].value == 42


@pytest.mark.django_db()
def test_decimal_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DecimalFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "decimal_field"
    ws["B2"].value = 4.2
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = DecimalFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.decimal_field == Decimal("4.2")


@pytest.mark.django_db()
def test_decimal_field_is_serialized() -> None:
    # Arrange.
    obj = DecimalFieldModel._default_manager.create(pk=1, decimal_field=Decimal("4.2"))

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.DecimalFieldModel"]["A1"].value == "id"
    assert wb["tests.DecimalFieldModel"]["A2"].value == 1
    assert wb["tests.DecimalFieldModel"]["B1"].value == "decimal_field"
    assert wb["tests.DecimalFieldModel"]["B2"].value == Decimal("4.2")


@pytest.mark.django_db()
def test_float_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.FloatFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "float_field"
    ws["B2"].value = 4.2
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = FloatFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.float_field == 4.2


@pytest.mark.django_db()
def test_float_field_is_serialized() -> None:
    # Arrange.
    obj = FloatFieldModel._default_manager.create(pk=1, float_field=4.2)

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.FloatFieldModel"]["A1"].value == "id"
    assert wb["tests.FloatFieldModel"]["A2"].value == 1
    assert wb["tests.FloatFieldModel"]["B1"].value == "float_field"
    assert wb["tests.FloatFieldModel"]["B2"].value == 4.2
