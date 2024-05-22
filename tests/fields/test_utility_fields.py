from __future__ import annotations

import uuid
from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import (
    EmailFieldModel,
    GenericIPAddressFieldModel,
    JSONFieldModel,
    URLFieldModel,
    UUIDFieldModel,
)

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db()
def test_email_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.EmailFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "email_field"
    ws["B2"].value = "92403542+paduszyk@users.noreply.github.com."
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = EmailFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.email_field == "92403542+paduszyk@users.noreply.github.com."


@pytest.mark.django_db()
def test_email_field_is_serialized() -> None:
    # Arrange.
    obj = EmailFieldModel._default_manager.create(
        pk=1,
        email_field="92403542+paduszyk@users.noreply.github.com.",
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.EmailFieldModel"]["A1"].value == "id"
    assert wb["tests.EmailFieldModel"]["A2"].value == 1
    assert wb["tests.EmailFieldModel"]["B1"].value == "email_field"
    assert wb["tests.EmailFieldModel"]["B2"].value == "92403542+paduszyk@users.noreply.github.com."  # fmt: skip


@pytest.mark.django_db()
def test_generic_ip_address_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("GenericIPAddressFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "generic_ip_address_field"
    ws["B2"].value = "127.0.0.1"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = GenericIPAddressFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.generic_ip_address_field == "127.0.0.1"


@pytest.mark.django_db()
def test_generic_ip_address_field_is_serialized() -> None:
    # Arrange.
    obj = GenericIPAddressFieldModel._default_manager.create(
        pk=1,
        generic_ip_address_field="127.0.0.1",
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["GenericIPAddressFieldModel"]["A1"].value == "id"
    assert wb["GenericIPAddressFieldModel"]["A2"].value == 1
    assert wb["GenericIPAddressFieldModel"]["B1"].value == "generic_ip_address_field"
    assert wb["GenericIPAddressFieldModel"]["B2"].value == "127.0.0.1"


@pytest.mark.django_db()
def test_json_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.JSONFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "json_field"
    ws["B2"].value = '{"key_1": 42, "key_2": [42], "key_3": {"4": 2}}'
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = JSONFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.json_field == {"key_1": 42, "key_2": [42], "key_3": {"4": 2}}


@pytest.mark.django_db()
def test_json_field_is_serialized() -> None:
    # Arrange.
    obj = JSONFieldModel._default_manager.create(
        pk=1,
        json_field={"key_1": 42, "key_2": [42], "key_3": {"4": 2}},
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.JSONFieldModel"]["A1"].value == "id"
    assert wb["tests.JSONFieldModel"]["A2"].value == 1
    assert wb["tests.JSONFieldModel"]["B1"].value == "json_field"
    assert wb["tests.JSONFieldModel"]["B2"].value == '{"key_1": 42, "key_2": [42], "key_3": {"4": 2}}'  # fmt: skip


@pytest.mark.django_db()
def test_url_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.URLFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "url_field"
    ws["B2"].value = "https://github.com/paduszyk/django-xlsx-serializer"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = URLFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.url_field == "https://github.com/paduszyk/django-xlsx-serializer"


@pytest.mark.django_db()
def test_url_field_is_serialized() -> None:
    # Arrange.
    obj = URLFieldModel._default_manager.create(
        pk=1,
        url_field="https://github.com/paduszyk/django-xlsx-serializer",
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.URLFieldModel"]["A1"].value == "id"
    assert wb["tests.URLFieldModel"]["A2"].value == 1
    assert wb["tests.URLFieldModel"]["B1"].value == "url_field"
    assert wb["tests.URLFieldModel"]["B2"].value == "https://github.com/paduszyk/django-xlsx-serializer"  # fmt: skip


@pytest.mark.django_db()
def test_uuid_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.UUIDFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "uuid_field"
    ws["B2"].value = "ebc54031-d5ae-4367-8033-e6fbaa3ca8d7"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = UUIDFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.uuid_field == uuid.UUID("ebc54031-d5ae-4367-8033-e6fbaa3ca8d7")


@pytest.mark.django_db()
def test_uuid_field_is_serialized() -> None:
    # Arrange.
    obj = UUIDFieldModel._default_manager.create(
        pk=1,
        uuid_field=uuid.UUID("ebc54031-d5ae-4367-8033-e6fbaa3ca8d7"),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.UUIDFieldModel"]["A1"].value == "id"
    assert wb["tests.UUIDFieldModel"]["A2"].value == 1
    assert wb["tests.UUIDFieldModel"]["B1"].value == "uuid_field"
    assert wb["tests.UUIDFieldModel"]["B2"].value == "ebc54031-d5ae-4367-8033-e6fbaa3ca8d7"  # fmt: skip
