from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import BooleanFieldModel

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.parametrize(
    "field_value",
    [
        False,
        True,
    ],
)
@pytest.mark.django_db
def test_boolean_field_is_deserialized(fixture_path: Path, field_value: bool) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.BooleanFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "boolean_field"
    ws["B2"].value = field_value
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = BooleanFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.boolean_field is field_value


@pytest.mark.parametrize(
    "field_value",
    [
        False,
        True,
    ],
)
@pytest.mark.django_db
def test_boolean_field_is_serialized(field_value: bool) -> None:
    # Arrange.
    obj = BooleanFieldModel._default_manager.create(pk=1, boolean_field=field_value)

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.BooleanFieldModel"]["A1"].value == "id"
    assert wb["tests.BooleanFieldModel"]["A2"].value == 1
    assert wb["tests.BooleanFieldModel"]["B1"].value == "boolean_field"
    assert wb["tests.BooleanFieldModel"]["B2"].value is field_value
