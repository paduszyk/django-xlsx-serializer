from __future__ import annotations

import datetime
from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize
from django.test.utils import override_settings

from tests.models import (
    DateFieldModel,
    DateTimeFieldModel,
    DurationFieldModel,
    TimeFieldModel,
)

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db()
def test_date_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DateFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "date_field"
    ws["B2"].value = "2024-04-02"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = DateFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.date_field == datetime.date(2024, 4, 2)


@pytest.mark.django_db()
def test_date_field_is_serialized() -> None:
    # Arrange.
    obj = DateFieldModel._default_manager.create(
        pk=1,
        date_field=datetime.date(2024, 4, 2),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.DateFieldModel"]["A1"].value == "id"
    assert wb["tests.DateFieldModel"]["A2"].value == 1
    assert wb["tests.DateFieldModel"]["B1"].value == "date_field"
    assert wb["tests.DateFieldModel"]["B2"].value == "2024-04-02"


@pytest.mark.django_db()
def test_time_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.TimeFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "time_field"
    ws["B2"].value = "22:44:42"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = TimeFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.time_field == datetime.time(22, 44, 42)


@pytest.mark.django_db()
def test_time_field_is_serialized() -> None:
    # Arrange.
    obj = TimeFieldModel._default_manager.create(
        pk=1,
        time_field=datetime.time(22, 44, 42),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.TimeFieldModel"]["A1"].value == "id"
    assert wb["tests.TimeFieldModel"]["A2"].value == 1
    assert wb["tests.TimeFieldModel"]["B1"].value == "time_field"
    assert wb["tests.TimeFieldModel"]["B2"].value == "22:44:42"


@override_settings(
    USE_TZ=False,
)
@pytest.mark.django_db()
def test_datetime_field_is_deserialized_with_tz_disabled(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DateTimeFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "datetime_field"
    ws["B2"].value = "2024-04-02T22:44:42"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = DateTimeFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.datetime_field == datetime.datetime(2024, 4, 2, 22, 44, 42)  # noqa: DTZ001


@override_settings(
    USE_TZ=True,
)
@pytest.mark.django_db()
def test_datetime_field_is_deserialized_with_tz_enabled(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DateTimeFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "datetime_field"
    ws["B2"].value = "2024-04-02T22:44:42+02:00"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = DateTimeFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.datetime_field == datetime.datetime(
        *(2024, 4, 2, 22, 44, 42),
        tzinfo=datetime.timezone(datetime.timedelta(hours=2)),
    )


@override_settings(
    USE_TZ=False,
)
@pytest.mark.django_db()
def test_datetime_field_is_serialized_with_tz_disabled() -> None:
    # Arrange.
    obj = DateTimeFieldModel._default_manager.create(
        pk=1,
        datetime_field=datetime.datetime(2024, 4, 2, 22, 44, 42),  # noqa: DTZ001
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.DateTimeFieldModel"]["A1"].value == "id"
    assert wb["tests.DateTimeFieldModel"]["A2"].value == 1
    assert wb["tests.DateTimeFieldModel"]["B1"].value == "datetime_field"
    assert wb["tests.DateTimeFieldModel"]["B2"].value == "2024-04-02T22:44:42"


@override_settings(
    USE_TZ=True,
)
@pytest.mark.django_db()
def test_datetime_field_is_serialized_with_tz_enabled() -> None:
    # Arrange.
    obj = DateTimeFieldModel._default_manager.create(
        pk=1,
        datetime_field=datetime.datetime(
            *(2024, 4, 2, 22, 44, 42),
            tzinfo=datetime.timezone(datetime.timedelta(hours=2)),
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.DateTimeFieldModel"]["A1"].value == "id"
    assert wb["tests.DateTimeFieldModel"]["A2"].value == 1
    assert wb["tests.DateTimeFieldModel"]["B1"].value == "datetime_field"
    assert wb["tests.DateTimeFieldModel"]["B2"].value == "2024-04-02T22:44:42+02:00"


@pytest.mark.django_db()
def test_duration_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.DurationFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "duration_field"
    ws["B2"].value = "04:02:42"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = DurationFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.duration_field == datetime.timedelta(hours=4, minutes=2, seconds=42)


@pytest.mark.django_db()
def test_duration_field_is_serialized() -> None:
    # Arrange.
    obj = DurationFieldModel._default_manager.create(
        pk=1,
        duration_field=datetime.timedelta(hours=4, minutes=2, seconds=42),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.DurationFieldModel"]["A1"].value == "id"
    assert wb["tests.DurationFieldModel"]["A2"].value == 1
    assert wb["tests.DurationFieldModel"]["B1"].value == "duration_field"
    assert wb["tests.DurationFieldModel"]["B2"].value == "04:02:42"
