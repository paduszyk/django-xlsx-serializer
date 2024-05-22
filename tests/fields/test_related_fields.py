from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import (
    ForeignKeyModel,
    ManyToManyFieldModel,
    NaturalKeyModel,
    OneToOneFieldModel,
    PrimaryKeyModel,
)

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db()
def test_foreign_key_field_is_deserialized_with_auto_fks(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "nk_field_1"
    ws["B2"].value = "value"
    ws["C1"].value = "nk_field_2"
    ws["C2"].value = 42
    ws = wb.create_sheet("tests.ForeignKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = 1
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = 1
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = ForeignKeyModel._default_manager.get(pk=1)
    related_primary_key_model_obj = PrimaryKeyModel._default_manager.get(pk=1)
    related_natural_key_model_obj = NaturalKeyModel._default_manager.get(pk=1)

    # Assert.
    assert obj.to_pk_model_field == related_primary_key_model_obj
    assert obj.to_nk_model_field == related_natural_key_model_obj


@pytest.mark.django_db()
def test_foreign_key_field_is_deserialized_with_natural_fks(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "nk_field_1"
    ws["A2"].value = "value"
    ws["B1"].value = "nk_field_2"
    ws["B2"].value = 42
    ws = wb.create_sheet("tests.ForeignKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = 1
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = "('value', 42)"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = ForeignKeyModel._default_manager.get(pk=1)
    related_primary_key_model_obj = PrimaryKeyModel._default_manager.get(pk=1)
    related_natural_key_model_obj = NaturalKeyModel._default_manager.get(
        nk_field_1="value",
        nk_field_2=42,
    )

    # Assert.
    assert obj.to_pk_model_field == related_primary_key_model_obj
    assert obj.to_nk_model_field == related_natural_key_model_obj


@pytest.mark.django_db()
def test_foreign_key_field_is_serialized_with_use_natural_fks_disabled() -> None:
    # Arrange
    obj = ForeignKeyModel._default_manager.create(
        pk=1,
        to_pk_model_field=PrimaryKeyModel._default_manager.create(pk=1),
        to_nk_model_field=NaturalKeyModel._default_manager.create(
            pk=2,
            nk_field_1="value",
            nk_field_2=42,
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=False)

    # Assert.
    assert wb["tests.ForeignKeyModel"]["A1"].value == "id"
    assert wb["tests.ForeignKeyModel"]["A2"].value == 1
    assert wb["tests.ForeignKeyModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.ForeignKeyModel"]["B2"].value == 1
    assert wb["tests.ForeignKeyModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.ForeignKeyModel"]["C2"].value == 2


@pytest.mark.django_db()
def test_foreign_key_field_is_serialized_with_use_natural_fks_enabled() -> None:
    # Arrange
    obj = ForeignKeyModel._default_manager.create(
        pk=1,
        to_pk_model_field=PrimaryKeyModel._default_manager.create(pk=1),
        to_nk_model_field=NaturalKeyModel._default_manager.create(
            nk_field_1="value",
            nk_field_2=42,
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=True)

    # Assert.
    assert wb["tests.ForeignKeyModel"]["A1"].value == "id"
    assert wb["tests.ForeignKeyModel"]["A2"].value == 1
    assert wb["tests.ForeignKeyModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.ForeignKeyModel"]["B2"].value == 1
    assert wb["tests.ForeignKeyModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.ForeignKeyModel"]["C2"].value == "('value', 42)"


@pytest.mark.django_db()
def test_one_to_one_field_is_deserialized_with_auto_fks(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "nk_field_1"
    ws["B2"].value = "value"
    ws["C1"].value = "nk_field_2"
    ws["C2"].value = 42
    ws = wb.create_sheet("tests.OneToOneFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = 1
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = 1
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = OneToOneFieldModel._default_manager.get(pk=1)
    related_primary_key_model_obj = PrimaryKeyModel._default_manager.get(pk=1)
    related_natural_key_model_obj = NaturalKeyModel._default_manager.get(pk=1)

    # Assert.
    assert obj.to_pk_model_field == related_primary_key_model_obj
    assert obj.to_nk_model_field == related_natural_key_model_obj


@pytest.mark.django_db()
def test_one_to_one_field_is_deserialized_with_natural_fks(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "nk_field_1"
    ws["A2"].value = "value"
    ws["B1"].value = "nk_field_2"
    ws["B2"].value = 42
    ws = wb.create_sheet("tests.OneToOneFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = 1
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = "('value', 42)"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = OneToOneFieldModel._default_manager.get(pk=1)
    related_primary_key_model_obj = PrimaryKeyModel._default_manager.get(pk=1)
    related_natural_key_model_obj = NaturalKeyModel._default_manager.get(
        nk_field_1="value",
        nk_field_2=42,
    )

    # Assert.
    assert obj.to_pk_model_field == related_primary_key_model_obj
    assert obj.to_nk_model_field == related_natural_key_model_obj


@pytest.mark.django_db()
def test_one_to_one_field_is_serialized_with_use_natural_fks_disabled() -> None:
    # Arrange
    obj = OneToOneFieldModel._default_manager.create(
        pk=1,
        to_pk_model_field=PrimaryKeyModel._default_manager.create(pk=1),
        to_nk_model_field=NaturalKeyModel._default_manager.create(
            pk=2,
            nk_field_1="value",
            nk_field_2=42,
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=False)

    # Assert.
    assert wb["tests.OneToOneFieldModel"]["A1"].value == "id"
    assert wb["tests.OneToOneFieldModel"]["A2"].value == 1
    assert wb["tests.OneToOneFieldModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.OneToOneFieldModel"]["B2"].value == 1
    assert wb["tests.OneToOneFieldModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.OneToOneFieldModel"]["C2"].value == 2


@pytest.mark.django_db()
def test_one_to_one_field_is_serialized_with_use_natural_foreign_keys_enabled() -> None:
    # Arrange
    obj = OneToOneFieldModel._default_manager.create(
        pk=1,
        to_pk_model_field=PrimaryKeyModel._default_manager.create(pk=1),
        to_nk_model_field=NaturalKeyModel._default_manager.create(
            nk_field_1="value",
            nk_field_2=42,
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=True)

    # Assert.
    assert wb["tests.OneToOneFieldModel"]["A1"].value == "id"
    assert wb["tests.OneToOneFieldModel"]["A2"].value == 1
    assert wb["tests.OneToOneFieldModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.OneToOneFieldModel"]["B2"].value == 1
    assert wb["tests.OneToOneFieldModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.OneToOneFieldModel"]["C2"].value == "('value', 42)"


@pytest.mark.django_db()
def test_many_to_many_field_is_deserialized_with_auto_fks(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["A3"].value = 2
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 3
    ws["A3"].value = 4
    ws["B1"].value = "nk_field_1"
    ws["B2"].value = "value"
    ws["B3"].value = "value"
    ws["C1"].value = "nk_field_2"
    ws["C2"].value = 42
    ws["C3"].value = 43
    ws = wb.create_sheet("tests.ManyToManyFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = "[1, 2]"
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = "[3, 4]"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = ManyToManyFieldModel._default_manager.get(pk=1)
    related_primary_key_model_obj_1 = PrimaryKeyModel._default_manager.get(pk=1)
    related_primary_key_model_obj_2 = PrimaryKeyModel._default_manager.get(pk=2)
    related_natural_key_model_obj_1 = NaturalKeyModel._default_manager.get(pk=3)
    related_natural_key_model_obj_2 = NaturalKeyModel._default_manager.get(pk=4)

    # Assert.
    assert list(obj.to_pk_model_field.all()) == [
        related_primary_key_model_obj_1,
        related_primary_key_model_obj_2,
    ]
    assert list(obj.to_nk_model_field.all()) == [
        related_natural_key_model_obj_1,
        related_natural_key_model_obj_2,
    ]


@pytest.mark.django_db()
def test_many_to_many_field_is_deserialized_with_natural_fks(fixture_path: Path) -> None:  # fmt: skip
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.PrimaryKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["A3"].value = 2
    ws = wb.create_sheet("tests.NaturalKeyModel")
    ws["A1"].value = "id"
    ws["A2"].value = 3
    ws["A3"].value = 4
    ws["B1"].value = "nk_field_1"
    ws["B2"].value = "value"
    ws["B3"].value = "value"
    ws["C1"].value = "nk_field_2"
    ws["C2"].value = 42
    ws["C3"].value = 43
    ws = wb.create_sheet("tests.ManyToManyFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "to_pk_model_field"
    ws["B2"].value = "[1, 2]"
    ws["C1"].value = "to_nk_model_field"
    ws["C2"].value = "[('value', 42), ('value', 43)]"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = ManyToManyFieldModel._default_manager.get(pk=1)
    related_primary_key_model_obj_1 = PrimaryKeyModel._default_manager.get(pk=1)
    related_primary_key_model_obj_2 = PrimaryKeyModel._default_manager.get(pk=2)
    related_natural_key_model_obj_1 = NaturalKeyModel._default_manager.get(pk=3)
    related_natural_key_model_obj_2 = NaturalKeyModel._default_manager.get(pk=4)

    # Assert.
    assert list(obj.to_pk_model_field.all()) == [
        related_primary_key_model_obj_1,
        related_primary_key_model_obj_2,
    ]
    assert list(obj.to_nk_model_field.all()) == [
        related_natural_key_model_obj_1,
        related_natural_key_model_obj_2,
    ]


@pytest.mark.django_db()
def test_many_to_many_field_is_serialized_with_use_natural_fks_disabled() -> None:
    # Arrange.
    obj = ManyToManyFieldModel._default_manager.create(pk=1)
    obj.to_pk_model_field.add(
        PrimaryKeyModel._default_manager.create(pk=1),
        PrimaryKeyModel._default_manager.create(pk=2),
    )
    obj.to_nk_model_field.add(
        NaturalKeyModel._default_manager.create(
            pk=3,
            nk_field_1="value",
            nk_field_2=42,
        ),
        NaturalKeyModel._default_manager.create(
            pk=4,
            nk_field_1="value",
            nk_field_2=43,
        ),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=False)

    # Assert.
    assert wb["tests.ManyToManyFieldModel"]["A1"].value == "id"
    assert wb["tests.ManyToManyFieldModel"]["A2"].value == 1
    assert wb["tests.ManyToManyFieldModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.ManyToManyFieldModel"]["B2"].value == "[1, 2]"
    assert wb["tests.ManyToManyFieldModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.ManyToManyFieldModel"]["C2"].value == "[3, 4]"


@pytest.mark.django_db()
def test_many_to_many_field_is_serialized_with_use_natural_fks_enabled() -> None:
    # Arrange.
    obj = ManyToManyFieldModel._default_manager.create(pk=1)
    obj.to_pk_model_field.add(
        PrimaryKeyModel._default_manager.create(pk=1),
        PrimaryKeyModel._default_manager.create(pk=2),
    )
    obj.to_nk_model_field.add(
        NaturalKeyModel._default_manager.create(nk_field_1="value", nk_field_2=42),
        NaturalKeyModel._default_manager.create(nk_field_1="value", nk_field_2=43),
    )

    # Act.
    wb = serialize("xlsx", [obj], use_natural_foreign_keys=True)

    # Assert.
    assert wb["tests.ManyToManyFieldModel"]["A1"].value == "id"
    assert wb["tests.ManyToManyFieldModel"]["A2"].value == 1
    assert wb["tests.ManyToManyFieldModel"]["B1"].value == "to_pk_model_field"
    assert wb["tests.ManyToManyFieldModel"]["B2"].value == "[1, 2]"
    assert wb["tests.ManyToManyFieldModel"]["C1"].value == "to_nk_model_field"
    assert wb["tests.ManyToManyFieldModel"]["C2"].value == "[('value', 42), ('value', 43)]"  # fmt: skip
