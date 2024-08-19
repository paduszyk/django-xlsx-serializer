from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import AutoFieldModel, BigAutoFieldModel, SmallAutoFieldModel

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db
def test_auto_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.AutoFieldModel")
    ws["A1"].value = "auto_field"
    ws["A2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = AutoFieldModel._default_manager.get(auto_field=42)

    # Assert.
    assert obj.auto_field == 42


@pytest.mark.django_db
def test_auto_field_is_serialized() -> None:
    # Arrange.
    obj = AutoFieldModel._default_manager.create()

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.AutoFieldModel"]["A1"].value == "auto_field"
    assert wb["tests.AutoFieldModel"]["A2"].value == obj.pk


@pytest.mark.django_db
def test_big_auto_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.BigAutoFieldModel")
    ws["A1"].value = "big_auto_field"
    ws["A2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = BigAutoFieldModel._default_manager.get(big_auto_field=42)

    # Assert.
    assert obj.big_auto_field == 42


@pytest.mark.django_db
def test_big_auto_field_is_serialized() -> None:
    # Arrange.
    obj = BigAutoFieldModel._default_manager.create()

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.BigAutoFieldModel"]["A1"].value == "big_auto_field"
    assert wb["tests.BigAutoFieldModel"]["A2"].value == obj.pk


@pytest.mark.django_db
def test_small_auto_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.SmallAutoFieldModel")
    ws["A1"].value = "small_auto_field"
    ws["A2"].value = 42
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = SmallAutoFieldModel._default_manager.get(small_auto_field=42)

    # Assert.
    assert obj.small_auto_field == 42


@pytest.mark.django_db
def test_small_auto_field_is_serialized() -> None:
    # Arrange.
    obj = SmallAutoFieldModel._default_manager.create()

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.SmallAutoFieldModel"]["A1"].value == "small_auto_field"
    assert wb["tests.SmallAutoFieldModel"]["A2"].value == obj.pk
