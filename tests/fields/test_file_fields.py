from __future__ import annotations

from typing import TYPE_CHECKING, Callable

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import FileFieldModel, FilePathFieldModel, ImageFieldModel

if TYPE_CHECKING:
    from pathlib import Path
    from unittest import mock


@pytest.mark.django_db
def test_file_field_is_deserialized(
    fixture_path: Path,
    mock_file: Callable[[str], mock.Mock],
) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.FileFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "file_field"
    ws["B2"].value = "file.txt"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = FileFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.file_field == mock_file("file.txt")


@pytest.mark.django_db
def test_file_field_is_serialized(mock_file: Callable[[str], mock.Mock]) -> None:
    # Arrange.
    obj = FileFieldModel._default_manager.create(pk=1, file_field=mock_file("file.txt"))

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.FileFieldModel"]["A1"].value == "id"
    assert wb["tests.FileFieldModel"]["A2"].value == 1
    assert wb["tests.FileFieldModel"]["B1"].value == "file_field"
    assert wb["tests.FileFieldModel"]["B2"].value == "file.txt"


@pytest.mark.django_db
def test_image_field_is_deserialized(
    fixture_path: Path,
    mock_image: Callable[[str], mock.Mock],
) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.ImageFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "image_field"
    ws["B2"].value = "image.png"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = ImageFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.image_field == mock_image("image.png")


@pytest.mark.django_db
def test_image_field_is_serialized(mock_image: Callable[[str], mock.Mock]) -> None:
    # Arrange.
    obj = ImageFieldModel._default_manager.create(
        pk=1,
        image_field=mock_image("image.png"),
    )

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.ImageFieldModel"]["A1"].value == "id"
    assert wb["tests.ImageFieldModel"]["A2"].value == 1
    assert wb["tests.ImageFieldModel"]["B1"].value == "image_field"
    assert wb["tests.ImageFieldModel"]["B2"].value == "image.png"


@pytest.mark.django_db
def test_file_path_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.FilePathFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "file_path_field"
    ws["B2"].value = "file.txt"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = FilePathFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.file_path_field == "file.txt"


@pytest.mark.django_db
def test_file_path_field_is_serialized() -> None:
    # Arrange.
    obj = FilePathFieldModel._default_manager.create(pk=1, file_path_field="file.txt")

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.FilePathFieldModel"]["A1"].value == "id"
    assert wb["tests.FilePathFieldModel"]["A2"].value == 1
    assert wb["tests.FilePathFieldModel"]["B1"].value == "file_path_field"
    assert wb["tests.FilePathFieldModel"]["B2"].value == "file.txt"
