from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

from django.core.management import call_command
from django.core.serializers import serialize

from tests.models import CharFieldModel, SlugFieldModel, TextFieldModel

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.django_db
def test_char_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.CharFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "char_field"
    ws["B2"].value = "value"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = CharFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.char_field == "value"


@pytest.mark.django_db
def test_char_field_is_serialized() -> None:
    # Arrange.
    obj = CharFieldModel._default_manager.create(pk=1, char_field="value")

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.CharFieldModel"]["A1"].value == "id"
    assert wb["tests.CharFieldModel"]["A2"].value == 1
    assert wb["tests.CharFieldModel"]["B1"].value == "char_field"
    assert wb["tests.CharFieldModel"]["B2"].value == "value"


@pytest.mark.django_db
def test_text_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.TextFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "text_field"
    ws["B2"].value = "value"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = TextFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.text_field == "value"


@pytest.mark.django_db
def test_text_field_is_serialized() -> None:
    # Arrange.
    obj = TextFieldModel._default_manager.create(pk=1, text_field="value")

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.TextFieldModel"]["A1"].value == "id"
    assert wb["tests.TextFieldModel"]["A2"].value == 1
    assert wb["tests.TextFieldModel"]["B1"].value == "text_field"
    assert wb["tests.TextFieldModel"]["B2"].value == "value"


@pytest.mark.django_db
def test_slug_field_is_deserialized(fixture_path: Path) -> None:
    # Arrange.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("tests.SlugFieldModel")
    ws["A1"].value = "id"
    ws["A2"].value = 1
    ws["B1"].value = "slug_field"
    ws["B2"].value = "value"
    wb.save(fixture_path)

    # Act.
    call_command("loaddata", fixture_path)
    obj = SlugFieldModel._default_manager.get(pk=1)

    # Assert.
    assert obj.slug_field == "value"


@pytest.mark.django_db
def test_slug_field_is_serialized() -> None:
    # Arrange.
    obj = SlugFieldModel._default_manager.create(pk=1, slug_field="value")

    # Act.
    wb = serialize("xlsx", [obj])

    # Assert.
    assert wb["tests.SlugFieldModel"]["A1"].value == "id"
    assert wb["tests.SlugFieldModel"]["A2"].value == 1
    assert wb["tests.SlugFieldModel"]["B1"].value == "slug_field"
    assert wb["tests.SlugFieldModel"]["B2"].value == "value"
