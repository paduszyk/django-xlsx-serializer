from __future__ import annotations

__all__ = [
    "Deserializer",
    "Serializer",
]

import ast
import io
import json
import sys
import warnings
from contextlib import suppress
from pathlib import Path
from typing import TYPE_CHECKING, Any, Final

import openpyxl

from django.apps import apps
from django.core.serializers import python
from django.core.serializers.base import DeserializedObject, SerializationError
from django.db import models

if TYPE_CHECKING:
    from collections.abc import Iterator

    from django.db.models import Model

if sys.version_info >= (3, 12):
    from typing import override
else:
    from typing_extensions import override

SHEET_NAME_MAX_LENGTH: Final[int] = 31

SHEET_NAME_INVALID_CHARACTERS: Final[str] = "\\?*:/[]"


def _get_model(model_identifier: str) -> type[Model]:
    # Determine the model's name and app label.
    if "." in (model_label := model_identifier.lower()):
        app_label, model_name = model_label.split(".")
    else:
        app_label, model_name = None, model_label

    # Determine the candidates for the model to be returned.
    models_candidates = [
        model
        for model in apps.get_models()
        if model._meta.model_name == model_name
        and (True if app_label is None else model._meta.app_label == app_label)
    ]

    # An identifier is valid if it represents a unique model.
    if (num_model_candidates := len(models_candidates)) != 1:
        if num_model_candidates == 0:
            msg = f"model {model_identifier!r} isn't installed"
        else:
            app_labels = ", ".join(
                [f"{model._meta.app_label!r}" for model in models_candidates],
            )
            msg = (
                f"model {model_identifier!r} found in multiple apps ({app_labels}); "
                f"use a fully qualified label to properly refer to the requested model"
            )
        raise LookupError(msg)

    return models_candidates[0]


class Serializer(python.Serializer):
    # Make the serializer discoverable with the `dumpdata` command.
    internal_use_only = False

    # If the stream is not specified in the options, force it to be set to `None`
    # (the base serializer class uses `io.StringIO` by default).
    stream_class = type(None)

    def get_model_sheet_names(self) -> dict[type[Model], str]:
        model_sheet_names_option = self.options.get("model_sheet_names", {})

        # Validate and reformat the `model_sheet_names` option value.
        # Reformatting replaces model identifiers with model classes.
        for model_identifier, sheet_name in model_sheet_names_option.copy().items():
            # Validate the model identifier.
            try:
                model = _get_model(model_identifier)
            except LookupError as e:
                msg = f"invalid 'model_sheet_names' option: {e}"

                raise SerializationError(msg) from e

            # Validate the sheet name.
            invalid_sheet_name_reasons: list[str] = []
            if (sheet_name_length := len(sheet_name)) > SHEET_NAME_MAX_LENGTH:
                invalid_sheet_name_reasons.append(
                    f"it is too long, {sheet_name_length} > {SHEET_NAME_MAX_LENGTH}",
                )
            if invalid_sheet_name_characters := [
                char for char in sheet_name if char in SHEET_NAME_INVALID_CHARACTERS
            ]:
                invalid_sheet_name_reasons.append(
                    f"it contains invalid characters: "
                    f"{', '.join(map(repr, invalid_sheet_name_characters))}",
                )
            if list(model_sheet_names_option.values()).count(sheet_name) > 1:
                invalid_sheet_name_reasons.append("it is not unique")

            if invalid_sheet_name_reasons:
                msg = (
                    f"{sheet_name!r} is not a valid Excel sheet name "
                    f"({'; '.join(invalid_sheet_name_reasons)})"
                )
                raise SerializationError(msg)

            model_sheet_names_option[model] = sheet_name

        model_sheet_names = {model: model._meta.label for model in apps.get_models()}
        model_sheet_names.update(model_sheet_names_option)

        return model_sheet_names

    @override
    def start_serialization(self) -> None:
        super().start_serialization()

        # Instantiate the output workbook.
        self._workbook = openpyxl.Workbook()

        # Models will be mapped into sheet names. The `model_sheet_names` option can be
        # passed to the `serialize()` method to specify custom model sheet names. The
        # value of this option should represent a mapping of model identifiers (either
        # fully qualified labels or names) to the desired sheet names.
        self._model_sheet_names = self.get_model_sheet_names()

        # Keep track of the sheet names added by the serializer.
        self._sheet_names_added: list[str] = []

    @override
    def end_object(self, obj: Any) -> None:
        super().end_object(obj)

        # Use the last Python object returned by the base serializer.
        obj = self.objects[-1]

        # Get the object's model to determine the default name for the sheet where the
        # object is going to be serialized.
        opts = obj["model"]._meta

        # Get sheet name corresponding to the model; by default, it's the model's label.
        sheet_name = self._model_sheet_names[opts.model]

        # Create/get & update the output sheet.
        if sheet_name not in self._workbook:
            if (model_sheet_name_length := len(sheet_name)) > SHEET_NAME_MAX_LENGTH:
                # This block can only be reached in the case of sheet names NOT passed
                # to the `model_sheet_names` option (as those were validated within the
                # `start_serialization()` method). The sheet name being checked here is
                # a fully qualified model label, and the only issue that might arise is
                # if it is too long.

                # Very long model labels are replaced by model names and then truncated
                # to the leading `SHEET_NAME_MAX_LENGTH` characters.
                sheet_name = sheet_name.split(".")[1][:SHEET_NAME_MAX_LENGTH]

                # An extra check for duplicate sheet names.
                if sheet_name in self._workbook:
                    msg = (
                        f"the truncated sheet name {sheet_name!r} for serializing the "
                        f"{opts.label!r} isn't unique; use the 'model_sheet_names' "
                        f"option to manually resolve too long or conflicting names"
                    )
                    raise SerializationError(msg)

                msg = (
                    f"{opts.label!r} objects are serialized into {sheet_name!r} sheet "
                    f"(fully qualified label is too long, {model_sheet_name_length} > "
                    f"{SHEET_NAME_MAX_LENGTH})"
                )
                warnings.warn(msg, RuntimeWarning, stacklevel=1)

            # Create the sheet and initialize it with the column headers.
            model_sheet = self._workbook.create_sheet(sheet_name)
            model_sheet.append(list(obj["fields"].keys()))

            # Update the `model_sheet_names` dict.
            self._model_sheet_names[opts.model] = sheet_name

            # Update the list of the sheet names added.
            self._sheet_names_added.append(sheet_name)
        else:
            model_sheet = self._workbook[sheet_name]

        # Serialize the object as another row.
        model_sheet.append(list(obj["fields"].values()))

    @override
    def end_serialization(self) -> None:
        super().end_serialization()

        # Remove sheets not added by the deserializer.
        for sheet_name in self._workbook.sheetnames.copy():
            if sheet_name not in self._sheet_names_added:
                del self._workbook[sheet_name]

        # Use the serializer's stream to determine the path for serialization output.
        # Based on the output stream type attempt to distinguish between a file path
        # (when called via the `serialize()` function from `django.core.serializers`)
        # and a stream object (when called via the `dumpdata` management command).

        stream = self.stream
        if isinstance(stream := self.stream, io.TextIOBase):
            output = None if (output := stream.name) == sys.stdout.name else output
            if output is None:
                msg = "printing workbooks to the standard output isn't supported"
                warnings.warn(msg, RuntimeWarning, stacklevel=1)
        else:
            if stream is not None and not isinstance(stream, (str, Path)):
                msg = "the stream must be a file path 'str' or 'pathlib.Path' object"
                raise SerializationError(msg)
            output = stream

        if output:
            if self.objects:
                self._workbook.save(output)
            else:
                msg = "the output workbook is empty, so it won't be saved"
                warnings.warn(msg, RuntimeWarning, stacklevel=1)

    @override
    def get_dump_object(self, obj: Model) -> dict[str, Any]:
        data = super().get_dump_object(obj)

        model = apps.get_model(data["model"])
        opts = model._meta

        # Update the object's field values prior to the final serialization.
        fields = data["fields"]
        for name, value in fields.items():
            field = opts.get_field(name)

            # Handle natural FKs, i.e. tuples assigned to `ForeignKey` fields.
            if (
                isinstance(field, models.ForeignKey)
                and value
                and isinstance(value, tuple)
            ):
                fields[name] = str(value)

            # Serialize the values of many-to-many fields as stringified Python objects
            # (regardless of whether they are serialized using natural keys or not).
            if isinstance(field, models.ManyToManyField):
                fields[name] = str(value)

            # Serialize `date`, `datetime`, and `time` objects as ISO 8601 strings.
            if (
                isinstance(
                    field,
                    (
                        models.DateField,
                        models.DateTimeField,
                        models.TimeField,
                    ),
                )
                and value
            ):
                fields[name] = value.isoformat()

            # Serialize `JSONField` values as plain strings.
            if isinstance(field, models.JSONField) and value:
                fields[name] = json.dumps(value, cls=field.encoder)

        # Update the `model` value.
        data["model"] = opts.model

        # Move PK to `fields` dict.
        data["fields"] = {
            **(
                {opts.pk.attname: pk}
                if (pk := data.pop("pk", None)) is not None
                else {}
            ),
            **data["fields"],
        }

        return data

    @override
    def getvalue(self) -> Any:
        return self._workbook


class Deserializer:
    def __init__(self, workbook_path: str | Path, **options: Any) -> None:
        # Load the workbook data.
        self._workbook = openpyxl.load_workbook(workbook_path)

        # Pass the options.
        self._options = options

    def __iter__(self) -> Iterator[DeserializedObject]:  # noqa: C901
        # Map models into the workbook's sheets.
        model_sheets: dict[type[Model], Any] = {}
        for sheet in self._workbook:
            # A model is identified based on the sheet name, which is supposed to be
            # either its fully qualified label or name (the latter applies only if the
            # model name is unique).
            with suppress(LookupError):
                model = _get_model(sheet.title)
                model_sheets[model] = sheet

        # Convert Excel data into Python objects.
        python_objects: list[dict[str, Any]] = []
        for model, sheet in model_sheets.items():
            opts = model._meta

            # Delete empty rows and columns.
            if (min_row := sheet.min_row) > 1:
                sheet.delete_rows(1, min_row - 1)
            if (min_column := sheet.min_column) > 1:
                sheet.delete_cols(1, min_column - 1)

            # Delete columns that don't represent any of the model's fields.
            sheet_columns = [cell.value for cell in sheet[1]]
            non_field_columns = [
                sheet_column
                for sheet_column in sheet_columns
                if sheet_column
                not in [
                    field.name for field in opts.local_fields + opts.local_many_to_many
                ]
            ]
            for non_field_column in non_field_columns:
                non_field_column_index = sheet_columns.index(non_field_column)
                sheet.delete_cols(non_field_column_index + 1)
                sheet_columns.remove(non_field_column)

            # Deserialize the sheet rows into the Python deserializer's format.
            python_model_objects = [
                {
                    "model": opts.label_lower,
                    "fields": dict(zip(sheet_columns, worksheet_row)),
                }
                for worksheet_row in sheet.iter_rows(
                    min_row=sheet.min_row + 1,
                    values_only=True,
                )
            ]
            python_objects += python_model_objects

        # Next, format Python objects for deserialization.
        for python_object in python_objects:
            model = apps.get_model(python_object["model"])

            fields = python_object["fields"]
            for name, value in fields.items():
                field = model._meta.get_field(name)

                # Handle valid data types representing an empty cell.
                if value in ("", None):
                    if field.null:
                        fields[name] = None
                    elif field.blank:
                        fields[name] = ""
                    # Continuing at this point may lead to integrity errors (tested).
                    continue

                # Handle natural foreign keys and many-to-many relations.
                if isinstance(
                    field,
                    (
                        models.ForeignKey,
                        models.ManyToManyField,
                        models.OneToOneField,
                    ),
                ) and isinstance(value, str):
                    fields[name] = ast.literal_eval(value)

                # Handle JSON values.
                if isinstance(field, models.JSONField):
                    fields[name] = json.loads(value, cls=field.decoder)

        return python.Deserializer(python_objects, **self._options)
